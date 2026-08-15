"""Local .xlsx snapshots. Not the store — see sheets.py for that.

The store is the live Google Sheet. This module exists for the moments when
you want the data as a file you can open, mail, or roll back to:

    store.backup()          before anything destructive
    ipopulse report         the formatted, human-readable workbook (report.py)

It writes the same tab layout `tables.py` defines, so a snapshot can be read
straight back with `read()` — which is what makes it a usable undo rather
than just a printout.
"""

from __future__ import annotations

import os
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

from . import tables


class WorkbookLocked(RuntimeError):
    """The file is held by another program — almost always Excel itself."""


def write(dest: Path, records: dict[str, dict]) -> Path:
    """Write records to `dest` as a workbook, atomically."""
    grid = tables.to_tables(records)

    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in grid.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row)
        _fit(ws)

    dest.parent.mkdir(parents=True, exist_ok=True)
    # Land atomically: a half-written zip is an unopenable workbook.
    fd, tmp = tempfile.mkstemp(dir=str(dest.parent), suffix=".xlsx.tmp")
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, dest)
    except PermissionError as exc:
        # Windows locks an open workbook. The raw error names a temp file
        # nobody recognises; say what actually has to happen instead.
        raise WorkbookLocked(
            f"{dest.name} is open in Excel (or another program is holding "
            f"it). Close it and run the command again — nothing was written."
        ) from exc
    finally:
        if os.path.exists(tmp):
            os.unlink(tmp)
    return dest


def read(path: Path) -> dict[str, dict]:
    """A snapshot back into records — the inverse of `write`."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        grid = {ws.title: [list(row) for row in ws.iter_rows(values_only=True)]
                for ws in wb.worksheets}
    finally:
        wb.close()
    return tables.from_tables(grid)


def _fit(ws) -> None:
    """Readable column widths — these files are meant to be opened."""
    from openpyxl.styles import Font
    widths: dict[int, int] = {}
    for row in ws.iter_rows(values_only=True):
        for i, cell in enumerate(row, 1):
            text = "" if cell is None else str(cell)
            widths[i] = min(60, max(widths.get(i, 10), len(text) + 2))
    for i, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True)
