"""Spreadsheet importer — Excel or CSV, from a local path or a URL.

Built to be tolerant, because the sheet you hand it will not match a schema
anyone agreed on. Column headers are matched fuzzily against the aliases below
(case, spaces, punctuation and the usual "(Rs Cr)" suffixes are ignored), and
anything unrecognised is reported rather than silently dropped.

Two shapes are supported:

  kind="ipos"  one row per IPO      -> creates/updates a YAML per row
  kind="gmp"   one row per GMP day  -> appends to that IPO's gmp_history

A Google Sheet works if you publish it as CSV
(File → Share → Publish to web → CSV), or export .xlsx and pass the path.
"""

from __future__ import annotations

import csv
import io
import os
import re
import tempfile
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

USER_AGENT = "ipo-pulse/1.0"

# canonical field -> header spellings seen in the wild
ALIASES: dict[str, list[str]] = {
    "company":     ["company", "companyname", "iponame", "name", "issuer", "ipo"],
    "board":       ["board", "type", "ipotype", "category", "segment"],
    "sector":      ["sector", "industry", "businesssector"],
    "fresh_cr":    ["freshissue", "fresh", "freshissuesize", "freshcr"],
    "ofs_cr":      ["ofs", "offerforsale", "ofssize", "ofscr"],
    "price_low":   ["pricelow", "lowerband", "minprice", "floorprice", "pricebandlow", "lower"],
    "price_high":  ["pricehigh", "upperband", "maxprice", "capprice", "pricebandhigh", "upper", "cutoffprice"],
    "lot_size":    ["lot", "lotsize", "marketlot", "minlot", "shareperlot"],
    "registrar":   ["registrar", "rta", "registraragent"],
    "announced":   ["announced", "announcedate", "announcementdate", "drhpdate"],
    "open":        ["open", "opendate", "issueopen", "openingdate", "startdate", "biddingstart"],
    "close":       ["close", "closedate", "issueclose", "closingdate", "enddate", "biddingend"],
    "allotment":   ["allotment", "allotmentdate", "basisofallotment"],
    "refund":      ["refund", "refunddate", "initiationofrefunds"],
    "listing":     ["listing", "listingdate", "listedon"],
    "eps":         ["eps", "earningspershare", "postissueeps"],
    "pe_peer_avg": ["peerpe", "industrype", "peeraveragepe", "sectorpe"],
    "revenue":     ["revenue", "totalrevenue", "sales", "turnover"],
    "ebitda":      ["ebitda", "operatingprofit", "ebidta"],
    "pat":         ["pat", "netprofit", "profitaftertax", "netincome"],
    "net_worth":   ["networth", "shareholdersfunds", "equity"],
    "total_debt":  ["totaldebt", "debt", "borrowings"],
    "gmp":         ["gmp", "greymarketpremium", "premium"],
    "kostak":      ["kostak", "kostakrate"],
    "sauda":       ["sauda", "subjecttosauda", "subject2sauda"],
    "date":        ["date", "ason", "asof", "gmpdate", "asondate"],
}

_norm_re = re.compile(r"[^a-z0-9]")


def _norm(text: Any) -> str:
    """'Price Band (High) ₹' -> 'pricebandhigh'."""
    s = str(text or "").lower()
    s = re.sub(r"\(.*?(rs|₹|cr|inr).*?\)", "", s)     # drop unit suffixes
    return _norm_re.sub("", s)


def build_header_map(headers: Iterable[Any]) -> tuple[dict[int, str], list[str]]:
    """Map column index -> canonical field. Also returns unmatched headers."""
    lookup: dict[str, str] = {}
    for field, names in ALIASES.items():
        for name in names:
            lookup[name] = field
        lookup[_norm(field)] = field

    mapping: dict[int, str] = {}
    unmatched: list[str] = []
    for i, raw in enumerate(headers):
        if raw in (None, ""):
            continue
        key = _norm(raw)
        if key in lookup:
            mapping[i] = lookup[key]
            continue
        hit = next((f for k, f in lookup.items() if k and (k in key or key in k)), None)
        if hit:
            mapping[i] = hit
        else:
            unmatched.append(str(raw))
    return mapping, unmatched


# ── value coercion ─────────────────────────────────────────────────────────

def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    s = re.sub(r"[₹,\s]", "", s)
    s = re.sub(r"(cr|crore|crores|rs|inr)$", "", s, flags=re.I)
    if s in ("", "-", "--", "na", "n/a", "nil"):
        return None
    neg = s.startswith("(") and s.endswith(")")          # (123) = -123
    if neg:
        s = s[1:-1]
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def to_date(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    if not s or s.lower() in ("na", "n/a", "-", "tba"):
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d %b %Y", "%d %B %Y",
                "%b %d, %Y", "%d-%b-%Y", "%d-%b-%y", "%Y/%m/%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else None


def slugify(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", str(text).lower()).strip("-")
    s = re.sub(r"-(ltd|limited|pvt|private)$", "", s)
    return s[:48] or "ipo"


# ── reading ────────────────────────────────────────────────────────────────

def _fetch(url: str) -> Path:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    suffix = ".xlsx" if ".xlsx" in url.lower() else ".csv"
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
    tmp = Path(tempfile.gettempdir()) / f"ipopulse-import{suffix}"
    tmp.write_bytes(data)
    return tmp


def read_rows(source: str, sheet: str | None = None) -> list[list[Any]]:
    """Return the sheet as a list of rows, header included."""
    path = _fetch(source) if source.startswith(("http://", "https://")) else Path(source)
    if not path.exists():
        raise FileNotFoundError(f"No such spreadsheet: {path}")

    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]

    raw = path.read_bytes().decode("utf-8-sig", errors="replace")
    dialect = csv.Sniffer().sniff(raw[:4096], delimiters=",;\t") if raw.strip() else csv.excel
    return [row for row in csv.reader(io.StringIO(raw), dialect)]


def find_header_row(rows: list[list[Any]], limit: int = 10) -> int:
    """Sheets often carry a title banner; pick the row that maps best."""
    best_i, best_n = 0, -1
    for i, row in enumerate(rows[:limit]):
        mapping, _ = build_header_map(row)
        if len(mapping) > best_n:
            best_i, best_n = i, len(mapping)
    return best_i


def parse(source: str, *, kind: str = "ipos", sheet: str | None = None) -> dict[str, Any]:
    """Read a spreadsheet into canonical records.

    Returns {records, unmatched, header_row, columns, skipped}.
    """
    rows = read_rows(source, sheet)
    if not rows:
        return {"records": [], "unmatched": [], "header_row": 0, "columns": {}, "skipped": 0}

    hrow = find_header_row(rows)
    mapping, unmatched = build_header_map(rows[hrow])

    records, skipped = [], 0
    for row in rows[hrow + 1:]:
        if not any(c not in (None, "") for c in row):
            continue
        rec: dict[str, Any] = {}
        for idx, field in mapping.items():
            if idx >= len(row):
                continue
            raw = row[idx]
            if field in ("company", "board", "sector", "registrar"):
                val = str(raw).strip() if raw not in (None, "") else None
            elif field in ("announced", "open", "close", "allotment", "refund", "listing", "date"):
                val = to_date(raw)
            else:
                val = to_number(raw)
            if val not in (None, ""):
                rec[field] = val
        # a GMP row needs a number; an IPO row needs a name
        needed = "gmp" if kind == "gmp" else "company"
        if needed not in rec:
            skipped += 1
            continue
        records.append(rec)

    return {"records": records, "unmatched": unmatched, "header_row": hrow,
            "columns": {v: k for k, v in mapping.items()}, "skipped": skipped}


# ── shaping into the canonical model ───────────────────────────────────────

def to_ipo_dict(rec: dict[str, Any], slug: str | None = None) -> dict[str, Any]:
    """One spreadsheet row -> a partial `Ipo` dict ready for merge()."""
    out: dict[str, Any] = {"slug": slug or slugify(rec.get("company", "ipo"))}
    for key in ("company", "board", "sector"):
        if rec.get(key):
            out[key] = rec[key]

    issue = {k: rec[k] for k in
             ("fresh_cr", "ofs_cr", "price_low", "price_high", "lot_size", "registrar")
             if rec.get(k) is not None}
    if issue:
        if "lot_size" in issue:
            issue["lot_size"] = int(issue["lot_size"])
        out["issue"] = issue

    dates = {k: rec[k] for k in ("announced", "open", "close", "allotment", "refund", "listing")
             if rec.get(k)}
    if dates:
        out["dates"] = dates

    fin = {k: [rec[k]] for k in ("revenue", "ebitda", "pat", "net_worth", "total_debt")
           if rec.get(k) is not None}
    if fin or rec.get("eps") or rec.get("pe_peer_avg"):
        fin["years"] = ["Latest"] if fin else []
        if rec.get("eps"):
            fin["eps"] = rec["eps"]
        if rec.get("pe_peer_avg"):
            fin["pe_peer_avg"] = rec["pe_peer_avg"]
        out["financials"] = fin

    if rec.get("gmp") is not None:
        out["gmp_history"] = [{
            "date": rec.get("date") or date.today().isoformat(),
            "gmp": rec["gmp"], "kostak": rec.get("kostak") or 0,
            "sauda": rec.get("sauda") or 0, "source": "sheet",
        }]
    return out


class SheetProvider:
    """Provider wrapper so `sync --provider sheet` works like the others."""

    name = "sheet"

    def __init__(self, source: str | None = None, sheet: str | None = None):
        self.source = source or os.getenv("IPOPULSE_SHEET_URL") or ""
        self.sheet = sheet

    def available(self) -> bool:
        return bool(self.source)

    def fetch_catalogue(self) -> list[dict[str, Any]]:
        parsed = parse(self.source, kind="ipos", sheet=self.sheet)
        return [to_ipo_dict(r) for r in parsed["records"]]

    def fetch_ipo(self, slug: str) -> dict[str, Any]:
        for rec in self.fetch_catalogue():
            if rec["slug"] == slug:
                return rec
        return {}

    def fetch_gmp(self, slug: str) -> list[dict[str, Any]]:
        parsed = parse(self.source, kind="gmp", sheet=self.sheet)
        out = []
        for rec in parsed["records"]:
            if rec.get("company") and slugify(rec["company"]) != slug:
                continue
            out.append({
                "date": rec.get("date") or date.today().isoformat(),
                "gmp": rec["gmp"], "kostak": rec.get("kostak") or 0,
                "sauda": rec.get("sauda") or 0, "source": "sheet",
            })
        return out

    def fetch_subscription(self, slug: str) -> list[dict[str, Any]]:
        return []
