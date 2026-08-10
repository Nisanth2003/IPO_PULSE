"""Excel report — every detail of an IPO in one workbook.

Sheets: Summary | Issue & Dates | Financials | GMP History | Subscription |
Analysis, plus an All IPOs board when you export the whole set.

Derived columns come from compute.py, so the spreadsheet and the on-screen
card can never disagree.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .compute import derive
from .models import Ipo

# ── house style ────────────────────────────────────────────────────────────
INK = "0F172A"
GREEN = "22C55E"
RED = "EF4444"

H1 = Font(bold=True, size=15, color="FFFFFF")
H2 = Font(bold=True, size=11, color="FFFFFF")
BOLD = Font(bold=True)
MUTED = Font(color="64748B", size=9)

FILL_HEAD = PatternFill("solid", fgColor=INK)
FILL_SUB = PatternFill("solid", fgColor="1E293B")
FILL_GREEN = PatternFill("solid", fgColor="DCFCE7")
FILL_RED = PatternFill("solid", fgColor="FEE2E2")
FILL_AMBER = PatternFill("solid", fgColor="FEF3C7")   # missing, but not blanking a scene
FILL_ALT = PatternFill("solid", fgColor="F8FAFC")

THIN = Side(style="thin", color="E2E8F0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

RUPEE = '"₹"#,##0.00'
RUPEE0 = '"₹"#,##0'
PCT = '0.0"%"'
MULT = '0.00"x"'


def _title(ws, text: str, span: int = 6) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=text)
    cell.font = H1
    cell.fill = FILL_HEAD
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def _header_row(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = H2
        cell.fill = FILL_SUB
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX


def _autosize(ws, minimum: int = 10, maximum: int = 46) -> None:
    for col in range(1, ws.max_column + 1):
        longest = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                longest = max(longest, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = max(
            minimum, min(maximum, longest + 3)
        )


def _kv(ws, row: int, key: str, value: Any, fmt: str | None = None) -> int:
    ws.cell(row=row, column=1, value=key).font = BOLD
    cell = ws.cell(row=row, column=2, value=value)
    if fmt:
        cell.number_format = fmt
    return row + 1


# ── sheets ─────────────────────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, ipo: Ipo, d: dict) -> None:
    ws = wb.create_sheet("Summary")
    _title(ws, f"{ipo.company or ipo.slug} — IPO Summary")
    r = 3
    r = _kv(ws, r, "Company", ipo.company)
    r = _kv(ws, r, "Board", ipo.board)
    r = _kv(ws, r, "Sector", ipo.sector)
    r = _kv(ws, r, "Status", d["dates"]["status"].title())
    r += 1
    r = _kv(ws, r, "Issue size (₹ Cr)", d["issue"]["total_cr"])
    r = _kv(ws, r, "Fresh issue (₹ Cr)", ipo.issue.fresh_cr)
    r = _kv(ws, r, "Fresh %", d["issue"]["fresh_pct"], PCT)
    r = _kv(ws, r, "OFS (₹ Cr)", ipo.issue.ofs_cr)
    r = _kv(ws, r, "OFS %", d["issue"]["ofs_pct"], PCT)
    r += 1
    r = _kv(ws, r, "Price band low", ipo.issue.price_low, RUPEE0)
    r = _kv(ws, r, "Price band high", ipo.issue.price_high, RUPEE0)
    r = _kv(ws, r, "Lot size", ipo.issue.lot_size)
    r = _kv(ws, r, "Minimum investment", d["issue"]["min_investment"], RUPEE0)
    r += 1
    g = d["gmp"]
    r = _kv(ws, r, "Latest GMP", g["gmp"], RUPEE0)
    r = _kv(ws, r, "GMP %", g["pct"], PCT)
    r = _kv(ws, r, "Estimated listing price", g["est_listing"], RUPEE0)
    r = _kv(ws, r, "Estimated gain per lot", g["gain_per_lot"], RUPEE0)
    r = _kv(ws, r, "Movement", g["movement"].title())
    r += 1
    s = d["subscription"]
    if s["has_data"]:
        r = _kv(ws, r, f"Subscription (Day {s['day']})", s["total"], MULT)
        r = _kv(ws, r, "  QIB", s["qib"], MULT)
        r = _kv(ws, r, "  NII / HNI", s["nii"], MULT)
        r = _kv(ws, r, "  Retail", s["retail"], MULT)
    r += 1
    sc = d["score"]
    r = _kv(ws, r, "IPO Pulse score", sc["effective"])
    r = _kv(ws, r, "  basis", f"{sc['source']}, scored on {sc['covered_pct']}% of the inputs")
    for part in sc["components"]:
        if part["has_data"]:
            r = _kv(ws, r, f"  {part['key']} ({part['weight']}%)",
                    f"{part['mark']}/10 — {part['detail']}")
    r = _kv(ws, r, "Verdict", ipo.analysis.verdict_text or ipo.analysis.verdict)

    ws.cell(row=r + 1, column=1,
            value="GMP is unofficial grey-market data. Not investment advice.").font = MUTED
    _autosize(ws)


def _sheet_issue(wb: Workbook, ipo: Ipo, d: dict) -> None:
    ws = wb.create_sheet("Issue & Dates")
    _title(ws, "Issue structure and key dates")
    r = 3
    ws.cell(row=r, column=1, value="ISSUE").font = BOLD
    r += 1
    for label, value, fmt in [
        ("Fresh issue (₹ Cr)", ipo.issue.fresh_cr, None),
        ("OFS (₹ Cr)", ipo.issue.ofs_cr, None),
        ("Total (₹ Cr)", d["issue"]["total_cr"], None),
        ("Price band", f"₹{ipo.issue.price_low:g} – ₹{ipo.issue.price_high:g}", None),
        ("Lot size", ipo.issue.lot_size, None),
        ("Minimum investment", d["issue"]["min_investment"], RUPEE0),
        ("Registrar", ipo.issue.registrar, None),
        ("Registrar URL", ipo.issue.registrar_url, None),
        ("Exchanges", ", ".join(ipo.issue.exchanges), None),
    ]:
        r = _kv(ws, r, label, value, fmt)

    r += 1
    ws.cell(row=r, column=1, value="DATES").font = BOLD
    r += 1
    dd = d["dates"]
    for label, key in [
        ("Announced", "announced"), ("Opens", "open"), ("Closes", "close"),
        ("Allotment", "allotment"), ("Refunds", "refund"), ("Listing", "listing"),
    ]:
        r = _kv(ws, r, label, dd.get(key) or "—")
    r = _kv(ws, r, "Cut-off time", ipo.dates.close_time)
    _autosize(ws)


def _sheet_financials(wb: Workbook, ipo: Ipo, d: dict) -> None:
    f = d["financials"]
    ws = wb.create_sheet("Financials")
    _title(ws, "Financials (₹ Cr unless stated)", span=10)
    if not f["has_data"]:
        ws.cell(row=3, column=1, value="No financial data entered yet.").font = MUTED
        _autosize(ws)
        return

    cols = ["Year", "Revenue", "EBITDA", "EBITDA %", "PAT", "PAT %",
            "Net worth", "Total debt", "RoNW %", "Debt/Equity"]
    _header_row(ws, 3, cols)
    for i, row in enumerate(f["rows"]):
        r = 4 + i
        vals = [row["year"], row["revenue"], row["ebitda"], row["ebitda_margin"],
                row["pat"], row["pat_margin"], row["net_worth"], row["total_debt"],
                row["ronw"], row["debt_equity"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BOX
            if i % 2:
                cell.fill = FILL_ALT
            if c in (4, 6, 9):
                cell.number_format = PCT
        ws.cell(row=r, column=1).font = BOLD

    r = 4 + len(f["rows"]) + 1
    ws.cell(row=r, column=1, value="DERIVED").font = BOLD
    r += 1
    for label, value, fmt in [
        ("Revenue CAGR", f["revenue_cagr"], PCT),
        ("EBITDA CAGR", f["ebitda_cagr"], PCT),
        ("PAT CAGR", f["pat_cagr"], PCT),
        ("EBITDA margin shift (bps)", f["margin_shift_bps"], None),
        ("EPS (post-issue)", f["eps"], None),
        ("P/E at upper band", f["pe"], None),
        ("Peer average P/E", f["pe_peer_avg"], None),
        ("Premium to peers", f["pe_premium_pct"], PCT),
        ("Market cap (₹ Cr)", f["market_cap_cr"], None),
    ]:
        r = _kv(ws, r, label, value, fmt)
    _autosize(ws)


def _sheet_gmp(wb: Workbook, ipo: Ipo, d: dict) -> None:
    ws = wb.create_sheet("GMP History")
    _title(ws, "Grey market premium — announcement to listing", span=6)
    _header_row(ws, 3, ["Date", "GMP (₹)", "GMP %", "Est. listing", "Kostak", "Source"])
    band = ipo.issue.price_high
    for i, p in enumerate(d["gmp"]["series"]):
        r = 4 + i
        vals = [p["date"], p["gmp"], p["pct"], round(band + p["gmp"], 2),
                p["kostak"], p["source"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BOX
            if c == 3:
                cell.number_format = PCT
                cell.fill = FILL_GREEN if p["gmp"] >= 0 else FILL_RED
            elif i % 2:
                cell.fill = FILL_ALT
    ws.freeze_panes = "A4"
    _autosize(ws)


def _sheet_subscription(wb: Workbook, ipo: Ipo, d: dict) -> None:
    ws = wb.create_sheet("Subscription")
    _title(ws, "Day-wise subscription (times)", span=7)
    _header_row(ws, 3, ["Day", "Date", "QIB", "NII / HNI", "Retail", "Employee", "Total"])
    for i, s in enumerate(d["subscription"].get("days", [])):
        r = 4 + i
        vals = [s["day"], s["date"], s["qib"], s["nii"], s["retail"],
                s["employee"], s["total"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BOX
            if c >= 3:
                cell.number_format = MULT
            if i % 2:
                cell.fill = FILL_ALT
        ws.cell(row=r, column=7).font = BOLD
    ws.freeze_panes = "A4"
    _autosize(ws)


def _sheet_analysis(wb: Workbook, ipo: Ipo, d: dict) -> None:
    ws = wb.create_sheet("Analysis")
    _title(ws, "Analysis and verdict", span=4)
    a = ipo.analysis
    r = 3
    ws.cell(row=r, column=1, value="BUSINESS OVERVIEW").font = BOLD
    r += 1
    for line in a.overview:
        ws.cell(row=r, column=1, value="•").alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=2, value=line)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="GREEN FLAGS").font = BOLD
    ws.cell(row=r, column=1).fill = FILL_GREEN
    r += 1
    for line in a.green_flags:
        ws.cell(row=r, column=2, value=line).fill = FILL_GREEN
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="RED FLAGS").font = BOLD
    ws.cell(row=r, column=1).fill = FILL_RED
    r += 1
    for line in a.red_flags:
        ws.cell(row=r, column=2, value=line).fill = FILL_RED
        r += 1

    r += 1
    for label, value in [
        ("Growth", a.growth), ("Valuation", a.valuation), ("Key risk", a.risk),
        ("Score (0-10)", d["score"]["effective"]),
        ("Verdict", a.verdict_text or a.verdict),
        ("Retail", a.reco_retail), ("HNI / NII", a.reco_hni), ("Long term", a.reco_long),
    ]:
        r = _kv(ws, r, label, value)
    _autosize(ws, maximum=70)


def _sheet_board(wb: Workbook, ipos: list[Ipo]) -> None:
    """All tracked IPOs side by side — the 'multiple IPOs at once' view."""
    ws = wb.create_sheet("All IPOs", 0)
    _title(ws, "All tracked IPOs", span=10)
    _header_row(ws, 3, ["Company", "Board", "Status", "Price band", "Lot",
                        "GMP ₹", "GMP %", "Est. listing", "Sub (x)", "Closes"])
    for i, ipo in enumerate(ipos):
        d = derive(ipo)
        r = 4 + i
        g, s = d["gmp"], d["subscription"]
        vals = [
            ipo.company or ipo.slug, ipo.board, d["dates"]["status"].title(),
            f"₹{ipo.issue.price_low:g}–₹{ipo.issue.price_high:g}",
            ipo.issue.lot_size, g["gmp"], g["pct"], g["est_listing"],
            s["total"] if s["has_data"] else "—",
            d["dates"]["close"] or "—",
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BOX
            if c == 7:
                cell.number_format = PCT
                cell.fill = FILL_GREEN if g["gmp"] >= 0 else FILL_RED
            elif i % 2:
                cell.fill = FILL_ALT
    ws.freeze_panes = "A4"
    _autosize(ws)


def _sheet_gaps(wb: Workbook, ipos: list[Ipo]) -> None:
    """What is missing, per IPO, and who can fill it.

    NSE supplies issue terms, dates and subscription; Gemini supplies GMP.
    Nothing free supplies financials — they live in the RHP PDF — so revenue,
    EBITDA, PAT and the valuation that depends on them stay zero until they
    are typed in. A zero renders on the reels as a confident "₹0", which is
    worse than a blank, so this sheet exists to make the holes findable
    before a scene is recorded rather than after.
    """
    from .doctor import CHECKS, FILLERS, gmp_gaps

    ws = wb.create_sheet("Data gaps", 1)
    _title(ws, "Missing data — fill these before recording", span=9)

    # Same list `ipopulse doctor` prints, so the sheet and the terminal can
    # never disagree about what counts as missing.
    _header_row(ws, 3, ["Field", "Breaks", "Filled by"]
                + [(i.company or i.slug)[:22] for i in ipos])
    for r, (label, getter, who, severity, breaks) in enumerate(CHECKS, start=4):
        ws.cell(row=r, column=1, value=label).border = BOX
        ws.cell(row=r, column=2, value=breaks).border = BOX
        ws.cell(row=r, column=3, value=FILLERS[who]).border = BOX
        for c, ipo in enumerate(ipos, start=4):
            try:
                ok = bool(getter(ipo))
            except Exception:
                ok = False
            cell = ws.cell(row=r, column=c,
                           value="ok" if ok else ("MISSING" if severity == "blank" else "—"))
            cell.border = BOX
            cell.alignment = Alignment(horizontal="center")
            cell.fill = FILL_GREEN if ok else (FILL_RED if severity == "blank" else FILL_AMBER)

    # A weekday with no reading makes reel 2's "daily" trail imply the premium
    # held steady when in fact nobody looked.
    gap_row = len(CHECKS) + 4
    ws.cell(row=gap_row, column=1, value="GMP trail gaps").border = BOX
    ws.cell(row=gap_row, column=2, value="reel 2 trail implies no movement").border = BOX
    ws.cell(row=gap_row, column=3, value="ipopulse gmp <slug> <value> --date <day>").border = BOX
    for c, ipo in enumerate(ipos, start=4):
        gaps = gmp_gaps(ipo)
        cell = ws.cell(row=gap_row, column=c, value="ok" if not gaps else f"{len(gaps)} day(s)")
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")
        cell.fill = FILL_GREEN if not gaps else FILL_AMBER

    note = ws.cell(row=len(CHECKS) + 6, column=1,
                   value="Financials are not published as data anywhere free — "
                         "they are in the RHP PDF. Type them into "
                         "backend/data/ipos/<slug>.yaml, then run: ipopulse build")
    note.alignment = Alignment(vertical="center")
    ws.freeze_panes = "D4"
    _autosize(ws)


# ── entry points ───────────────────────────────────────────────────────────

def build_workbook(ipos: list[Ipo], *, board: bool = False) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)                       # drop the default empty sheet
    for ipo in ipos:
        d = derive(ipo)
        if len(ipos) == 1:
            _sheet_summary(wb, ipo, d)
            _sheet_issue(wb, ipo, d)
            _sheet_financials(wb, ipo, d)
            _sheet_gmp(wb, ipo, d)
            _sheet_subscription(wb, ipo, d)
            _sheet_analysis(wb, ipo, d)
    if board or len(ipos) > 1:
        _sheet_board(wb, ipos)
    if ipos:
        _sheet_gaps(wb, ipos)
    if not wb.sheetnames:                      # nothing to write
        wb.create_sheet("Empty")
    return wb


def write_report(ipos: list[Ipo], path: Path, *, board: bool = False) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(ipos, board=board)
    wb.properties.creator = "IPO Pulse"
    wb.properties.created = datetime.now()
    wb.save(path)
    return path
